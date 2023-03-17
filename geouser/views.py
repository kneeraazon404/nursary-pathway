import json
from copy import deepcopy
from datetime import datetime

import django_filters
import requests
from config import settings
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.db.models import Case, Value, When
from django.db.models.aggregates import Sum
from django.db.models.expressions import F, Value
from django.db.models.functions import Concat
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Alignment
from openpyxl.styles.borders import BORDER_MEDIUM, BORDER_THIN, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.mixins import ListModelMixin
from rest_framework.permissions import AllowAny, IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet, ModelViewSet
from rest_framework_api_key.permissions import HasAPIKey
from rolepermissions.roles import get_user_roles
from user.tasks import send_user_activation
from utils import constants
from utils.constants import UserType
from utils.filters import GeoUserFilter, PriceRangeFilter, ProjectFilter
from utils.logger import logger
from utils.pagination import CustomPageNumberPagination
from utils.serializers import EmptySerializer, TotalSerializer

from .models import (
    GeoItems,
    GeokrishiCategory,
    GeokrishiProduct,
    GeokrishiSubCategory,
    GeokrishiUsers,
)
from .serializers import (
    GeoItemSerializer,
    GeokrishiCategorySerializer,
    GeokrishiProductSummarySerializer,
    GeokrishiSubCategorySerializer,
    GeoProductListSerializer,
    GeoProductSerializer,
    GeoUserSerializer,
    UserSerializer,
)


def geo_user(geotoken):
    r = requests.get(
        f"{settings.GEOKRISHI_BASE}/api/user/profile",
        headers={"Authorization": geotoken},
    )
    r_data = r.json()
    r_data["c_id"] = r_data.pop("id")
    id_value = r_data["c_id"]
    r_data["geo_user_id"] = r_data.pop("c_id")
    geo = GeokrishiUsers.objects.filter(geo_user_id=id_value)
    if geo:
        print("User already exists")
    else:
        print("No user")
        serializer = GeoUserSerializer(data=r_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
    return


class GeokrishiProxyView(APIView):
    """
    Proxy viewset to redirect to geokrishi apis
    """

    permission_classes = [AllowAny]

    def get_geotoken(self):
        geotoken = self.request.headers.get("GEOTOKEN", None)
        return geotoken

    def get_headers(self):
        headers = {}
        geotoken = self.get_geotoken()
        headers["AUTHORIZATION"] = geotoken
        return headers

    def get_url(self, path):
        query = "&".join([f"{k}={v}" for k, v in self.request.query_params.items()])
        if query != "":
            query = f"?{query}"
        return f"{settings.GEOKRISHI_BASE}/api/samuhik_bazar/{path}{query}"

    def get(self, request, path, format=None):
        data = None
        r = requests.get(
            self.get_url(path),
            headers=self.get_headers(),
        )

        if r.status_code >= 200 and r.status_code < 300:
            data = r.json()
        elif r.status_code == 400:
            data = r.json()
        else:
            data = r.content

        return Response(data, status=r.status_code)

    def post(self, request, path, format=None):
        r = requests.post(
            self.get_url(path),
            request.data,
            headers=self.get_headers(),
        )
        data = None

        if r.status_code >= 200 and r.status_code < 300:
            data = r.json()
        elif r.status_code == 400:
            data = r.json()
        else:
            data = r.content
        return Response(data, status=r.status_code)

    def put(self, request, path, format=None):
        r = requests.put(
            self.get_url(path),
            request.data,
            headers=self.get_headers(),
        )
        data = None

        if r.status_code >= 200 and r.status_code < 300:
            data = r.json()
        elif r.status_code == 400:
            data = r.json()
        else:
            data = r.content
        return Response(data, status=r.status_code)

    def patch(self, request, path, format=None):
        r = requests.patch(
            self.get_url(path),
            request.data,
            headers=self.get_headers(),
        )
        data = None

        if r.status_code >= 200 and r.status_code < 300:
            data = r.json()
        elif r.status_code == 400:
            data = r.json()
        else:
            data = r.content
        return Response(data, status=r.status_code)

    def delete(self, request, path, format=None):
        r = requests.delete(
            self.get_url(path),
            headers=self.get_headers(),
        )
        data = None

        if r.status_code >= 200 and r.status_code < 300:
            data = r.json()
        elif r.status_code == 400:
            data = r.json()
        else:
            data = r.content
        return Response(data, status=r.status_code)


class GeoUser(GenericViewSet, ListModelMixin):
    """
    ! Deprecated: API & Data migrated to geokrishi
    Import user details from geokrishi.
    """

    permission_classes = [HasAPIKey]
    serializer_class = GeoUserSerializer
    queryset = GeokrishiUsers.objects.all().order_by(
        "firstName",
        "lastName",
        "geo_user_id",
    )
    filter_backends = [
        DjangoFilterBackend,
    ]
    filterset_class = GeoUserFilter

    def create(self, request, *args, **kwargs):
        geotoken = request.headers.get("GEOTOKEN")
        print(geotoken)
        r = requests.get(
            f"{settings.GEOKRISHI_BASE}/api/user/profile",
            headers={
                # "Authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzQ1MDM4NzY1LCJqdGkiOiJkMTY3OWQ1MzIwZjI0ODZkYjhmMGY2NWFhNzFkZmM5NSIsInVzZXJfaWQiOjUyMX0.UY3Dng6p_84VrcEcqvr0P2BRAhSvgN8o2pyuQKxEoO0"
                "Authorization": geotoken
            },
        )
        r_data = r.json()
        r_data["c_id"] = r_data.pop("id")
        id_value = r_data["c_id"]
        r_data["geo_user_id"] = r_data.pop("c_id")
        geo = GeokrishiUsers.objects.filter(geo_user_id=id_value)
        if geo:
            print("User already exists")
        else:
            serializer = self.get_serializer(data=r_data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
        return Response(r.json())
        # return Response(r.json(), status=STATUS.HTTP_201_CREATED)

    def perform_create(self, serializer):
        serializer.save()


class GeoProduct(ModelViewSet):
    """
    ! Deprecated: API & Data migrated to geokrishi
    """

    permission_classes = [HasAPIKey | IsAuthenticatedOrReadOnly]
    serializer_class = GeoProductSerializer
    pagination_class = CustomPageNumberPagination
    filter_backends = [
        ProjectFilter,
        PriceRangeFilter,
        DjangoFilterBackend,
        # filters.SearchFilter,
        # filters.OrderingFilter,
    ]
    filterset_fields = [
        "items",
        "status",
        # "title__category",
        "available_date_eng",
        "geo_user__district",
        "geo_user__municipality",
        "geo_user",
    ]
    lookup_field = "id"
    # search_fields = ["price"]
    # ordering_fields = ["price"]

    def get_serializer_class(self):
        if (
            self.action == "except_product"
            or self.action == "my_product"
            or self.action == "expired_product"
            or self.action == "list"
        ):
            return GeoProductListSerializer
        # elif self.action=="update":
        #     return GeoProductUpdateSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        request = self.request
        status_order = [
            GeokrishiProduct.AVAILABLE,
            GeokrishiProduct.EXPIRED,
            GeokrishiProduct.CANCELED,
            GeokrishiProduct.SOLD,
        ]
        if self.action == "except_product":
            queryset = (
                GeokrishiProduct.objects.not_deleted()
                .exclude(geo_user=request.geouser)
                .order_by(
                    Case(
                        *[
                            When(status__iexact=siz, then=Value(i))
                            for i, siz in enumerate(status_order)
                        ],
                        default=None,
                    ),
                    "available_date_eng",
                )
            )
        elif self.action == "my_product":
            queryset = (
                GeokrishiProduct.objects.not_deleted()
                .filter(geo_user=request.geouser)
                .order_by(
                    Case(
                        *[
                            When(status__iexact=siz, then=Value(i))
                            for i, siz in enumerate(status_order)
                        ],
                        default=None,
                    ),
                    "available_date_eng",
                )
            )
        elif self.action == "expired_product":
            # instance=GeokrishiProduct.objects.not_deleted().filter(geo_user=request.geouser)
            # if instance.is_expired==True:
            #     queryset= instance
            # queryset= None
            queryset = (
                GeokrishiProduct.objects.not_deleted()
                .filter(
                    geo_user=request.geouser, available_date_eng__lt=datetime.today()
                )
                .order_by(
                    Case(
                        *[
                            When(status__iexact=siz, then=Value(i))
                            for i, siz in enumerate(status_order)
                        ],
                        default=None,
                    ),
                    "available_date_eng",
                )
            )  # this means available date is less than today
        else:
            queryset = GeokrishiProduct.objects.not_deleted().order_by(
                Case(
                    *[
                        When(status__iexact=siz, then=Value(i))
                        for i, siz in enumerate(status_order)
                    ],
                    default=None,
                ),
                "-available_date_eng",
            )

            # geo_user=GeokrishiUsers.objects.get(geo_user_id=request.geouser)
            # if geo_user.municipality:
            #     queryset=GeokrishiProduct.objects.not_deleted().filter(geo_user__municipality=geo_user.municipality)
            # elif geo_user.district:
            #     queryset=GeokrishiProduct.objects.not_deleted().filter(geo_user__district=geo_user.district)
            # elif geo_user.state:
            #     queryset=GeokrishiProduct.objects.not_deleted().filter(geo_user__state=geo_user.state)
            # else:
            #     queryset=GeokrishiProduct.objects.not_deleted()

        return queryset

    @action(
        detail=False,
        methods=["get"],
        url_path="except",
    )
    def except_product(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @action(
        detail=False,
        methods=["get"],
        url_path="expired",
    )
    def expired_product(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @action(
        methods=["GET"],
        detail=False,
        url_path=r"sold_report/(?P<from_date>[\d\-]+)/(?P<to_date>[\d\-]+)",
    )
    def sold_report(self, request, from_date, to_date, *args, **kwargs):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Sales Report.xlsx"'

        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(status="sold").filter(
            sold_date__date__gte=from_date, sold_date__date__lte=to_date
        )

        rows = queryset.annotate(
            name=Concat("geo_user__firstName", Value(" "), "geo_user__lastName"),
            address=Concat(
                "geo_user__district",
                Value(", "),
                "geo_user__municipality",
                Value(", "),
                "geo_user__settlement_name",
            ),
            # qty=Concat("quantity", Value(" "), "items__unit"),
            # total_price=F("selling_price") * F("quantity"),
        ).values_list(
            "name",
            "geo_user__district",
            "geo_user__municipality",
            "geo_user__settlement_name",
            "geo_user__wardNo",
            "geo_user__contactNo",
            "items__name",
            "quantity",
            "items__unit",
            "selling_price",
            # "total_price",
            "available_date_eng",
            "sold_date__date",
        )

        data = list()
        for index, row in enumerate(rows):
            row = list(row)
            row.insert(
                10,
                f"={get_column_letter(8)}{index+3}*{get_column_letter(10)}{index+3}",
            )
            data.append(row)

        title = [f"Product Sales from {from_date} to {to_date}"]
        headers = [
            "Farmer",
            "District",
            "Municipality",
            "Settlement Name",
            "Ward No",
            "Contact No",
            "Crop Name",
            "Quantity",
            "Unit",
            "Selling Price",
            "Total Price",
            "Available Date",
            "Sold Date",
        ]
        total_footer = [
            "Total",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            f"=SUM({get_column_letter(11)}3:{get_column_letter(11)}{len(data)+2})",
            "",
            "",
        ]

        total_columns = len(headers)
        title_rows = 2
        data_rows = len(data)
        total_rows = data_rows + title_rows + 1

        wb = Workbook()
        ws = wb.active
        ws.title = "Product Sales"

        ws.append(title)
        ws.append(headers)

        title_font = deepcopy(DEFAULT_FONT)
        title_font.b = True
        title_font.size = 16

        header_font = deepcopy(DEFAULT_FONT)
        header_font.b = True

        for row in data:
            ws.append(list(row))

        ws.append(total_footer)

        thin_top_border = Border(
            top=Side(style=BORDER_MEDIUM),
            left=Side(style=BORDER_THIN),
            right=Side(style=BORDER_THIN),
            bottom=Side(style=BORDER_THIN),
        )

        cell_border = Border(
            left=Side(style=BORDER_THIN),
            right=Side(style=BORDER_THIN),
            top=Side(style=BORDER_THIN),
            bottom=Side(style=BORDER_THIN),
        )

        # Data Row Borders
        for rw in range(title_rows - 1, total_rows + 1):
            for cl in range(1, total_columns + 1):
                cell = ws[f"{get_column_letter(cl)}{rw}"]
                # cell.number_format = "#,##0"
                cell.border = cell_border
                if rw == 1:
                    cell.font = title_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif rw <= title_rows:
                    cell.font = header_font

        ws.merge_cells("A1:" + get_column_letter(len(headers)) + "1")
        row = ws.row_dimensions[1]
        row.height = 25
        row.font = title_font
        row.alignment = Alignment(horizontal="center", vertical="center")

        # Total Row style
        for rw in range(data_rows + title_rows + 1, total_rows + 1):
            for i in range(1, total_columns + 1):
                cell = ws[f"{get_column_letter(i)}{rw}"]
                if rw == data_rows + title_rows + 1:
                    cell.border = thin_top_border

        dims = {}
        for index, row in enumerate(ws.rows):
            if index > 0:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value)))
                        )
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

        wb.save(response)
        return response

    @action(
        detail=False,
        methods=["get"],
        url_path="user/product",
    )
    def my_product(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @action(
        detail=False,
        methods=["get"],
        serializer_class=GeokrishiProductSummarySerializer,
    )
    def summary(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = (
            queryset.order_by("items__name")
            .values("items__name", "items__unit")
            .annotate(
                item_name=F("items__name"),
                item_unit=F("items__unit"),
                total_quantity=Sum("quantity"),
            )
        )
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer, request)
        return Response(
            {"data": serializer.data, "message": constants.PRODUCT_ADDED},
            status=status.HTTP_200_OK,
        )

    def perform_create(self, serializer, request):
        geouser = GeokrishiUsers.objects.get(geo_user_id=request.geouser)
        serializer.save(geo_user=geouser)
        # serializer.save(geo_user=request.geouser)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # if instance.is_expired==True:
        self.perform_destroy(instance)
        return Response(
            {"message": constants.PRODUCT_DELETED}, status=status.HTTP_200_OK
        )
        # return Response({"message":"Expired Items cannot be deleted"},status=status.HTTP_400_BAD_REQUEST)

    def perform_destroy(self, instance):
        instance.soft_delete()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, "_prefetched_objects_cache", None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        initial_status = serializer.initial_data.get("status")
        if initial_status == GeokrishiProduct.SOLD:
            instance.sold_date = timezone.now()
            instance.save()

        status = serializer.data.get("status")
        if status == GeokrishiProduct.EXPIRED:
            return Response({"message": constants.PRODUCT_EXPIRED})
        elif status == GeokrishiProduct.SOLD:
            return Response({"message": constants.PRODUCT_SOLD})
        elif status == GeokrishiProduct.CANCELED:
            return Response({"message": constants.PRODUCT_CANCELED})
        else:
            return Response({"message": constants.PRODUCT_ADDED})


class Acess(GenericViewSet):
    """
    ! Deprecated: API & Data migrated to geokrishi
    create user in samuhik with the given username,password and give acces to geokrishi
    """

    serializer_class = UserSerializer

    def perform_create(self, serializer):
        return serializer.save()

    def create(self, request, *args, **kwargs):
        username_ = request.data.get("username")
        password_ = request.data.get("password")
        ##create user/trader with that username and password
        if not username_:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            user = self.perform_create(serializer)
            user.is_active = False
            user.save()

        r = requests.post(
            settings.GEOKRISHI_URL,
            # "http://localhost:8000/o/token/",
            # "http://geokrishi.farm/o/token/",
            data={
                "grant_type": "password",
                "username": username_,
                "password": password_,
                "client_id": "YzdwgUwzhhY3L2gj6OzIZl55epfwVlmxzqoARJEz",
                "client_secret": "3m8qgcHfikfS2d0WH8s7LUAVhfF4QMFL4hQN7q4wkSHJNpW6ptLGMtTJzH4md1xfEiaE30wgVJDjgrYIQiNu0AzIwvYnSueVKd6oUeHbUnjDcsqZBMpZQM6zBdHU5CNC",
            },
        )
        status_code = r.status_code
        if not status_code == 200:
            raise Exception("Error accessing to samuhik, please try again")
        response = r.text
        response_json = r.json()
        return Response(response_json)


class GeokrishiCategoryViewset(ModelViewSet):

    """
    ! Deprecated: API & Data migrated to geokrishi
    Services of the agrovet or category of the Items items / products.
    """

    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = GeokrishiCategorySerializer
    queryset = GeokrishiCategory.objects.all()
    # queryset = GeokrishiCategory.objects.select_related("parent").filter(parent=None)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        print(queryset)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class GeokrishiSubCategory(ModelViewSet):

    """
    ! Deprecated: API & Data migrated to geokrishi
    Services of the agrovet or category of the Items items / products.
    """

    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = GeokrishiSubCategorySerializer
    queryset = GeokrishiSubCategory.objects.all()
    filter_backends = [
        PriceRangeFilter,
        DjangoFilterBackend,
        # filters.SearchFilter,
        # filters.OrderingFilter,
    ]
    filterset_fields = [
        "category",
    ]

    # def get_queryset(self):
    #     for i in GeokrishiSubCategory.objects.not_deleted():
    #         if i.name_set__exists():
    #             return i
    #     return None

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class GeoItemsViewset(ModelViewSet):

    """
    ! Deprecated: API & Data migrated to geokrishi
    Services of the agrovet or category of the Items items / products.
    """

    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = GeoItemSerializer
    queryset = GeoItems.objects.all()
    filter_backends = [
        DjangoFilterBackend,
        # filters.SearchFilter,
        # filters.OrderingFilter,
    ]
    filterset_fields = [
        "sub_category",
    ]
